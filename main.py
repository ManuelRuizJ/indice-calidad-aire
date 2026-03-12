import pandas as pd
import numpy as np
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

# ============================================================================
# CARGA DE CONFIGURACION DESDE JSON
# ============================================================================
with open('config.json', 'r', encoding='utf-8') as f:
    config = json.load(f)

# NADF-009
VENTANAS_NADF = config['NADF']['ventanas']
BANDAS_NADF = {k: [tuple(v) for v in val] for k, val in config['NADF']['bandas'].items()}
COLORES_NADF = {tuple(map(int, k.split('-'))): v for k, v in config['NADF']['colores'].items()}

# NOM-172
VENTANAS_NOM = config['NOM']['ventanas']
BANDAS_NOM = {k: [tuple(v) for v in val] for k, val in config['NOM']['bandas'].items()}
COLORES_NOM = config['NOM']['colores']

# Otros parametros
SUFICIENCIA = config['suficiencia']
ORDEN_CATEGORIAS = config['orden_categorias']

# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================

def calcular_ica(conc, bandas):
    for pcinf, pcsup, iinf, isup in bandas:
        if pcinf <= conc <= pcsup:
            k = (isup - iinf) / (pcsup - pcinf)
            return round((k * (conc - pcinf)) + iinf)
    return np.nan

def clasificar_nom(conc, bandas):
    if pd.isna(conc):
        return None
    for lim_inf, lim_sup, cat in bandas:
        if lim_inf < conc <= lim_sup:
            return cat
        elif conc == lim_inf and lim_inf == 0:
            return cat
    return None

def promedio_movil_simple(serie, ventana):
    min_datos = int(np.ceil(ventana * SUFICIENCIA))
    return serie.rolling(window=ventana, min_periods=min_datos).mean()

def nowcast(serie, pollutant):
    fa = 0.714 if pollutant == "PM10" else 0.694
    valores = serie.values
    n = len(valores)
    resultado = np.full(n, np.nan)
    for i in range(n):
        if i < 11:
            continue
        ultimas3 = valores[i-2:i+1]
        if np.sum(~np.isnan(ultimas3)) < 2:
            continue
        inicio = i - 11
        ventana = valores[inicio:i+1]
        validos = ventana[~np.isnan(ventana)]
        if len(validos) == 0:
            continue
        cmax = np.max(validos)
        cmin = np.min(validos)
        if cmax == 0:
            w = 1.0
        else:
            w = 1 - (cmax - cmin) / cmax
        W = round(max(w, 0.5), 2)
        suma_num = 0.0
        suma_den = 0.0
        for j, idx in enumerate(range(i, inicio-1, -1)):
            if j >= 12:
                break
            if not np.isnan(valores[idx]):
                peso = W ** j
                suma_num += valores[idx] * peso
                suma_den += peso
        if suma_den > 0:
            resultado[i] = (suma_num / suma_den) * fa
    return pd.Series(resultado, index=serie.index)

def redondear_nom(valor, contaminante, unidad):
    if pd.isna(valor):
        return np.nan
    if contaminante in ["PM10", "PM2.5"]:
        return int(round(valor))
    elif contaminante in ["O3", "NO2", "SO2"]:
        return round(valor, 3)
    elif contaminante == "CO":
        return round(valor, 2)
    else:
        return valor

def obtener_color_ica(valor):
    for (lo, hi), color in COLORES_NADF.items():
        if lo <= valor <= hi:
            return color
    return None

def preparar_datos_hoja(df):
    """Convierte las primeras filas en metadatos y reindexa a frecuencia horaria."""
    estaciones = df.iloc[0].values
    contaminantes = df.iloc[1].values
    unidades = df.iloc[2].values
    datos_raw = df.iloc[3:].reset_index(drop=True)

    dates_raw = datos_raw.iloc[:, 0]
    dates = pd.to_datetime(dates_raw, errors='coerce', dayfirst=True)
    if dates.isna().any():
        print(f"ADVERTENCIA: {dates.isna().sum()} filas con fecha no valida seran descartadas.")
        datos_raw = datos_raw.loc[~dates.isna()]
        dates = dates[~dates.isna()]
    datos_raw.index = dates
    datos_raw.index.name = 'Fecha'
    datos_raw = datos_raw.drop(columns=0)
    if datos_raw.index.duplicated().any():
        print(f"ADVERTENCIA: {datos_raw.index.duplicated().sum()} indices duplicados; se conserva la primera ocurrencia.")
        datos_raw = datos_raw[~datos_raw.index.duplicated(keep='first')]
    full_range = pd.date_range(start=datos_raw.index.min(), end=datos_raw.index.max(), freq='h')
    data_df = datos_raw.reindex(full_range)
    gaps = full_range.difference(data_df.index)
    if len(gaps) > 0:
        print(f"INFO: Se agregaron {len(gaps)} horas faltantes (NaN).")
    return estaciones, contaminantes, unidades, data_df, len(df.columns)

def peor_categoria(series_categorias, umbral_suficiencia=0.75):
    """
    Dadas varias series de categorias, devuelve la peor (mayor riesgo) por fila,
    solo si al menos el `umbral_suficiencia` de las series tienen datos no nulos.
    En caso contrario, retorna None.
    """
    if not series_categorias:
        return pd.Series(index=pd.Index([]), dtype='object')
    df_cat = pd.concat(series_categorias, axis=1)
    # Contar cuantas columnas tienen datos no nulos por fila
    count_valid = df_cat.notna().sum(axis=1)
    min_requerido = int(np.ceil(len(series_categorias) * umbral_suficiencia))
    # Convertir a valores numericos
    df_num = df_cat.apply(lambda col: col.map(ORDEN_CATEGORIAS).fillna(-1))
    max_num = df_num.max(axis=1)
    # Solo mantener donde hay suficiencia
    max_num = max_num.where(count_valid >= min_requerido, -1)
    inverso = {v: k for k, v in ORDEN_CATEGORIAS.items()}
    return max_num.map(inverso).where(max_num >= 0, None)

def combinar_con_existente(df_nuevo, archivo, nombre_hoja, col_fecha):
    """
    Lee la hoja 'nombre_hoja' del archivo existente (si existe) y la combina con df_nuevo.
    Retorna el DataFrame combinado (con indice de fecha).
    """
    if os.path.exists(archivo):
        try:
            df_existente = pd.read_excel(archivo, sheet_name=nombre_hoja, index_col=col_fecha)
            df_combinado = pd.concat([df_existente, df_nuevo], axis=0, sort=False)
            df_combinado = df_combinado[~df_combinado.index.duplicated(keep='last')]
            df_combinado.sort_index(inplace=True)
            return df_combinado
        except ValueError:
            # La hoja no existe, solo devolvemos el nuevo
            return df_nuevo
    else:
        return df_nuevo

def extraer_estaciones(df, tipo):
    """
    A partir de un DataFrame consolidado (con columnas como 'ICA_O3_Estacion' o 'AIRE_O3_Estacion'),
    devuelve un diccionario {nombre_estacion: DataFrame_con_columnas_de_esa_estacion}.
    Incluye la columna de fecha (indice). La columna 'Calidad del aire' se recalcula por estacion.
    """
    # Obtener lista de estaciones unicas
    estaciones = set()
    for col in df.columns:
        if tipo == 'ICA' and col.startswith('ICA_'):
            partes = col.split('_')
            if len(partes) >= 3:
                est = '_'.join(partes[2:])
                estaciones.add(est)
        elif tipo in ['AIRE', 'DIARIO'] and col.startswith('AIRE_'):
            partes = col.split('_')
            if len(partes) >= 3:
                est = '_'.join(partes[2:])
                estaciones.add(est)

    dfs_estacion = {}
    for est in sorted(estaciones):
        # Columnas de categoria y cantidad que terminan con esta estacion
        cols_cat = [c for c in df.columns if c.endswith(est) and 'CANTIDAD' not in c and c.startswith('AIRE_')]
        cols_cant = [c for c in df.columns if c.endswith(est) and 'CANTIDAD' in c]
        # Crear DataFrame con esas columnas
        df_est = df[cols_cat + cols_cant].copy()
        # Calcular calidad del aire para esta estacion
        if cols_cat:
            series_cat = [df_est[col] for col in cols_cat]
            df_est['Calidad del aire'] = peor_categoria(series_cat, SUFICIENCIA)
        dfs_estacion[est] = df_est
    return dfs_estacion

def aplicar_formato_ica(ws):
    """Aplica formato y colores a una hoja de ICA."""
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column > 1 and isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                color = obtener_color_ica(int(cell.value))
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 25

def aplicar_formato_aire(ws):
    """Aplica formato y colores a una hoja de AIRE Y SALUD (horario o diario)."""
    # Identificar columnas de categoria (empiezan con 'AIRE_' y no son 'CANTIDAD_')
    columnas_categoria = []
    for col in ws.iter_cols(min_row=1, max_row=1):
        if col[0].value and isinstance(col[0].value, str):
            if col[0].value.startswith('AIRE_') and 'CANTIDAD' not in col[0].value:
                columnas_categoria.append(col[0].column)
            elif col[0].value == 'Calidad del aire':
                columnas_categoria.append(col[0].column)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Colorear celdas de categoria
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in columnas_categoria and cell.value in COLORES_NOM:
                cell.fill = PatternFill(start_color=COLORES_NOM[cell.value],
                                        end_color=COLORES_NOM[cell.value],
                                        fill_type='solid')
                if cell.value in ['Buena', 'Aceptable']:
                    cell.font = Font(bold=True, color='000000')
                else:
                    cell.font = Font(bold=True, color='FFFFFF')

    # Formato de numero para columnas de cantidad
    for col in ws.columns:
        if col[0].value and isinstance(col[0].value, str) and col[0].value.startswith('CANTIDAD_'):
            nombre = col[0].value
            if 'O3' in nombre or 'NO2' in nombre or 'SO2' in nombre:
                fmt = '0.000'
            elif 'CO' in nombre:
                fmt = '0.00'
            else:
                fmt = '0'
            for cell in col[1:]:
                if cell.value is not None:
                    cell.number_format = fmt

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 25

def guardar_diccionario_excel(archivo, diccionario_dfs, tipo):
    """
    Guarda un diccionario de DataFrames en un archivo Excel, cada uno en una hoja.
    tipo: 'ICA', 'AIRE' o 'DIARIO' para aplicar el formato adecuado.
    """
    with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
        for nombre_hoja, df in diccionario_dfs.items():
            # Limitar nombre de hoja a 31 caracteres
            nombre_hoja = nombre_hoja[:31]
            df.to_excel(writer, sheet_name=nombre_hoja, index=True)

    # Aplicar formato a cada hoja
    wb = load_workbook(archivo)
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        if tipo == 'ICA':
            aplicar_formato_ica(ws)
        elif tipo == 'AIRE' or tipo == 'DIARIO':
            aplicar_formato_aire(ws)
    wb.save(archivo)

# ============================================================================
# PROCESAMIENTO PRINCIPAL
# ============================================================================
archivo_entrada = "datos/datos_calidad_aire.xlsx"
salida_ica = "datos/datos_calidad_aire_ICA.xlsx"
salida_aire = "datos/datos_calidad_aire_AIRE_Y_SALUD.xlsx"
salida_diario = "datos/datos_calidad_aire_DIARIO.xlsx"

xls = pd.ExcelFile(archivo_entrada)

# ----------------------------------------------------------------------------
# 1. Generar datos para ICA (NADF-009) en un DataFrame consolidado
# ----------------------------------------------------------------------------
print("Procesando datos para ICA...")
df_ica_total = pd.DataFrame()

for hoja in xls.sheet_names:
    df = pd.read_excel(xls, sheet_name=hoja, header=None)
    estaciones, contaminantes, unidades, data_df, num_orig_cols = preparar_datos_hoja(df)

    df_hoja = pd.DataFrame(index=data_df.index)

    for i in range(1, num_orig_cols):
        col_in_data = i - 1
        estacion = estaciones[i]
        contaminante = contaminantes[i]
        unidad = unidades[i]

        if not isinstance(contaminante, str) or contaminante == "Status":
            continue

        clave_orig = f"{contaminante}_{unidad}"
        if clave_orig not in VENTANAS_NADF:
            continue

        ventana = VENTANAS_NADF[clave_orig]
        valores = pd.to_numeric(data_df.iloc[:, col_in_data], errors="coerce")

        if i + 1 < num_orig_cols:
            status_series = data_df.iloc[:, i]
            status_str = status_series.astype(str).str.strip().str.lower()
            valores = valores.where(status_str == "ok", np.nan)

        valores = valores.where(valores >= 0, np.nan)

        if (valores == 0).all():
            print(f"ADVERTENCIA: {contaminante} en {estacion} tiene todos los valores en 0.")
            valores[:] = np.nan

        valores_prom = promedio_movil_simple(valores, ventana)

        if contaminante in ["O3", "NO2", "SO2"]:
            valores_prom = valores_prom / 1000.0
            clave_bandas = f"{contaminante}_ppm"
        else:
            clave_bandas = clave_orig

        ica_lista = [calcular_ica(x, BANDAS_NADF[clave_bandas]) if not np.isnan(x) else np.nan for x in valores_prom]
        df_hoja[f"ICA_{contaminante}_{estacion}"] = ica_lista

    df_hoja = df_hoja.dropna(how='all')
    if not df_hoja.empty:
        df_ica_total = pd.concat([df_ica_total, df_hoja], axis=0)

# Combinar con existente (hoja General)
df_ica_general = combinar_con_existente(df_ica_total, salida_ica, 'General', 'Fecha & Hora')
# Crear diccionario de hojas: General + estaciones (para ICA no hay columna de calidad)
diccionario_ica = {'General': df_ica_general}
# Para ICA no se calcula calidad del aire, asi que extraemos sin ese paso
# Usamos una version simplificada de extraer_estaciones sin calcular calidad
estaciones_ica = set()
for col in df_ica_general.columns:
    if col.startswith('ICA_'):
        partes = col.split('_')
        if len(partes) >= 3:
            est = '_'.join(partes[2:])
            estaciones_ica.add(est)
for est in sorted(estaciones_ica):
    cols_est = [c for c in df_ica_general.columns if c.endswith(est)]
    diccionario_ica[est] = df_ica_general[cols_est].copy()

guardar_diccionario_excel(salida_ica, diccionario_ica, 'ICA')
print("Archivo ICA generado/actualizado con hojas por estacion.")

# ----------------------------------------------------------------------------
# 2. Generar datos para AIRE Y SALUD horario (NOM-172)
# ----------------------------------------------------------------------------
print("Procesando datos para AIRE Y SALUD horario...")
df_aire_total = pd.DataFrame()

for hoja in xls.sheet_names:
    df = pd.read_excel(xls, sheet_name=hoja, header=None)
    estaciones, contaminantes, unidades, data_df, num_orig_cols = preparar_datos_hoja(df)

    df_hoja = pd.DataFrame(index=data_df.index)

    for i in range(1, num_orig_cols):
        col_in_data = i - 1
        estacion = estaciones[i]
        contaminante = contaminantes[i]
        unidad = unidades[i]

        if not isinstance(contaminante, str) or contaminante == "Status":
            continue

        clave_orig = f"{contaminante}_{unidad}"
        if clave_orig not in VENTANAS_NOM:
            continue

        valores = pd.to_numeric(data_df.iloc[:, col_in_data], errors="coerce")

        if i + 1 < num_orig_cols:
            status_series = data_df.iloc[:, i]
            status_str = status_series.astype(str).str.strip().str.lower()
            valores = valores.where(status_str == "ok", np.nan)

        valores = valores.where(valores >= 0, np.nan)

        if (valores == 0).all():
            print(f"ADVERTENCIA: {contaminante} en {estacion} tiene todos los valores en 0.")
            valores[:] = np.nan

        if clave_orig in ["PM10_ug/m3", "PM2.5_ug/m3"]:
            tipo = "PM10" if clave_orig.startswith("PM10") else "PM2.5"
            conc_base = nowcast(valores, tipo)
            clave_bandas = clave_orig
        elif clave_orig == "CO_ppm":
            conc_base = promedio_movil_simple(valores, 8)
            clave_bandas = "CO_ppm"
        else:
            conc_base = valores / 1000.0
            clave_bandas = f"{contaminante}_ppm"

        conc_redondeada = [redondear_nom(x, contaminante, unidad) for x in conc_base]
        categorias = [clasificar_nom(x, BANDAS_NOM[clave_bandas]) for x in conc_redondeada]

        col_cat = f"AIRE_{contaminante}_{estacion}"
        col_conc = f"CANTIDAD_{contaminante}_{estacion}"
        df_hoja[col_cat] = categorias
        df_hoja[col_conc] = conc_redondeada

    df_hoja = df_hoja.dropna(how='all')
    if not df_hoja.empty:
        df_aire_total = pd.concat([df_aire_total, df_hoja], axis=0)

# Calcular calidad global para la hoja General (sin umbral, con todas las categorias)
if not df_aire_total.empty:
    cols_cat = [c for c in df_aire_total.columns if c.startswith('AIRE_') and 'CANTIDAD' not in c]
    if cols_cat:
        series_cat = [df_aire_total[col] for col in cols_cat]
        # Para la hoja General, no aplicamos umbral (o aplicamos uno muy bajo)
        # Usamos la misma funcion pero con umbral 0 para que siempre calcule
        df_aire_total['Calidad del aire'] = peor_categoria(series_cat, 0.0)

# Combinar con existente
df_aire_general = combinar_con_existente(df_aire_total, salida_aire, 'General', 'Fecha & Hora')
# Extraer hojas por estacion (con su propia calidad)
diccionario_aire = {'General': df_aire_general}
diccionario_aire.update(extraer_estaciones(df_aire_general, 'AIRE'))
guardar_diccionario_excel(salida_aire, diccionario_aire, 'AIRE')
print("Archivo AIRE Y SALUD horario generado/actualizado con hojas por estacion.")

# ----------------------------------------------------------------------------
# 3. Generar datos para DIARIO (NOM-172)
# ----------------------------------------------------------------------------
print("Procesando datos para DIARIO...")
df_diario_total = pd.DataFrame()

for hoja in xls.sheet_names:
    df = pd.read_excel(xls, sheet_name=hoja, header=None)
    estaciones, contaminantes, unidades, data_df, num_orig_cols = preparar_datos_hoja(df)

    data_df['Fecha_dia'] = data_df.index.date
    dias = data_df['Fecha_dia'].unique()
    dias_ordenados = sorted(dias)
    df_dia = pd.DataFrame(index=dias_ordenados)

    for i in range(1, num_orig_cols):
        col_in_data = i - 1
        estacion = estaciones[i]
        contaminante = contaminantes[i]
        unidad = unidades[i]

        if not isinstance(contaminante, str) or contaminante == "Status":
            continue

        clave_orig = f"{contaminante}_{unidad}"
        if clave_orig not in VENTANAS_NOM:
            continue

        valores = pd.to_numeric(data_df.iloc[:, col_in_data], errors="coerce")

        if i + 1 < num_orig_cols:
            status_series = data_df.iloc[:, i]
            status_str = status_series.astype(str).str.strip().str.lower()
            valores = valores.where(status_str == "ok", np.nan)

        valores = valores.where(valores >= 0, np.nan)

        serie_valores = pd.Series(valores.values, index=data_df.index)

        if clave_orig in ["PM10_ug/m3", "PM2.5_ug/m3"]:
            min_datos = int(np.ceil(24 * SUFICIENCIA))
            valor_diario = serie_valores.resample('D').apply(
                lambda x: x.mean() if x.count() >= min_datos else np.nan
            )
            valor_redondeado = valor_diario.apply(lambda v: int(round(v)) if not pd.isna(v) else np.nan)
            clave_bandas = clave_orig
        elif clave_orig == "CO_ppm":
            prom_8h = promedio_movil_simple(serie_valores, 8)
            valor_diario = prom_8h.resample('D').max()
            valor_redondeado = valor_diario.apply(lambda v: round(v, 2) if not pd.isna(v) else np.nan)
            clave_bandas = "CO_ppm"
        else:
            serie_ppm = serie_valores / 1000.0
            valor_diario = serie_ppm.resample('D').max()
            valor_redondeado = valor_diario.apply(lambda v: round(v, 3) if not pd.isna(v) else np.nan)
            clave_bandas = f"{contaminante}_ppm"

        categorias = [clasificar_nom(v, BANDAS_NOM[clave_bandas]) for v in valor_redondeado]

        col_cat = f"AIRE_{contaminante}_{estacion}"
        col_conc = f"CANTIDAD_{contaminante}_{estacion}"
        df_dia[col_cat] = categorias
        df_dia[col_conc] = valor_redondeado.values

    df_dia = df_dia.dropna(how='all')
    if not df_dia.empty:
        df_diario_total = pd.concat([df_diario_total, df_dia], axis=0)

# Calcular calidad global diaria para hoja General
if not df_diario_total.empty:
    cols_cat = [c for c in df_diario_total.columns if c.startswith('AIRE_') and 'CANTIDAD' not in c]
    if cols_cat:
        series_cat = [df_diario_total[col] for col in cols_cat]
        df_diario_total['Calidad del aire'] = peor_categoria(series_cat, 0.0)

# Combinar con existente
df_diario_general = combinar_con_existente(df_diario_total, salida_diario, 'General', 'Fecha')
diccionario_diario = {'General': df_diario_general}
diccionario_diario.update(extraer_estaciones(df_diario_general, 'DIARIO'))
guardar_diccionario_excel(salida_diario, diccionario_diario, 'DIARIO')
print("Archivo DIARIO generado/actualizado con hojas por estacion.")